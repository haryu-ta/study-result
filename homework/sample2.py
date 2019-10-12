# -*- coding: utf-8 -*-

#=============================================================-
#
#  指定月月末日までのカレンダーを作成します
#
#=============================================================-

import datetime
import openpyxl
import calendar
from openpyxl import Workbook
from openpyxl.styles import Font,PatternFill,Border,Alignment
from openpyxl.styles.borders import Border,Side
from openpyxl.styles.colors import RED,BLUE,Color

#定義情報
#===================================================================================================
#指定年
year__c=2015
#指定月
month__c=9


#カラム幅定義
#diccolwidth = { "B" : 4.5 , "C" : 8.38 , "D" : 31.25 , "E"  : 29.50 , "F" : 33.88 }
diccolwidth = { "B" : 5.12 , "C" : 9.01 , "D" : 31.87 , "E"  : 30.12 , "F" : 34.51 }

#カラム使用列上限
maxcol = 6

#出力先ディレクトリ
outputdir="C:/temp"
#出力先ファイル名
outputfile="calender.xlsx"
#===================================================================================================

def sheetCreate(ws,yyyymm):
    #シート名変更
    ws.title = yyyymm.strftime('%Y%m')
    #ヘッダー作成
    ws.cell(row = 4,column=2).value = str(int(yyyymm.strftime('%m'))) + "月"
    ws.cell(row = 4,column=2).font = Font(size=18,bold=True,name='MS ゴシック')
    ws.cell(row = 4,column=2).fill = PatternFill(fill_type='solid',start_color='ffaa00',end_color='ffaa00')
    ws.cell(row = 4,column=2).alignment = Alignment(horizontal='center')
    ws.merge_cells(start_row=4,start_column=2,end_row=4,end_column=maxcol)


def createBorder(ws,rowind):
    for indx in range(2,maxcol + 1):
        for indy in range(4,rowind):
            ws.cell(row=indy,column=indx).border = Border(left=Side(border_style='thin',color='00000000'),
                     right=Side(border_style='thin',color='00000000'),
                     top=Side(border_style='thin',color='00000000'),
                     bottom=Side(border_style='thin',color='00000000'))


def addjustWidth(ws):
    for k,v in diccolwidth.items():
        ws.column_dimensions[k].width = v

def holidayColorling(ws,rowno,color):
    for indx in range(2,maxcol + 1):
        ws.cell(row=rowno,column=indx).fill = PatternFill(fill_type='solid',start_color=color,end_color=color)

def convertWeekday(weekday):
    if weekday == 0:
        return '月'
    elif weekday == 1:
        return '火'
    elif weekday == 2:
        return '水'
    elif  weekday == 3:
        return '木'
    elif  weekday == 4:
        return '金'
    elif  weekday == 5:
        return '土'
    else:
        return '日'

if __name__ == '__main__':

    rowind = 5

    #本日日付
    nowtime = datetime.date.today()
    #終了日付(指定月の翌月1日)
    totime = datetime.date(year__c,month__c,calendar.monthrange(year__c,month__c)[1]) + datetime.timedelta(days=1)
    rangetime = totime - nowtime

    #新規ワークブック
    wb = Workbook()
    ws = wb.get_active_sheet()
    sheetCreate(ws,(nowtime + datetime.timedelta(days=0)))

    for cnt in range(rangetime.days):

        standarttime = (nowtime + datetime.timedelta(days=cnt)).strftime('%Y%m')

        if not (standarttime in wb.get_sheet_names()):
            createBorder(ws,rowind)  #罫線作成
            addjustWidth(ws)         #列幅調整
            ws = wb.create_sheet()
            sheetCreate(ws,(nowtime + datetime.timedelta(days=cnt)))
            rowind = 5

        #日付セット
        ws.cell(row = rowind,column=2).value = str(int((nowtime + datetime.timedelta(days=cnt)).strftime('%d')))
        #曜日セット
        ws.cell(row = rowind,column=3).value = convertWeekday((nowtime + datetime.timedelta(days=cnt)).weekday())

        #土日は色付け
        weekdays = convertWeekday((nowtime + datetime.timedelta(days=cnt)).weekday())
        if weekdays == "土":
            holidayColorling(ws,rowind,Color(BLUE))
        elif weekdays == "日":
            holidayColorling(ws,rowind,Color(RED))

        #変数インクリメント
        rowind = rowind + 1

    #最終シートの罫線用
    createBorder(ws,rowind)
    #最終シート用列幅調整
    addjustWidth(ws)

    wb.save(outputdir + "/" + outputfile)